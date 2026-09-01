from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock, patch
import openpyxl
import requests

from ingestion.validation.url_validator import (
    LovdataURLValidator,
    URLAuditRecord,
    extract_lovdata_url,
    probe_url,
)


def test_extract_lovdata_url_law():
    doc = {"dok_id": "NL/lov/2005-06-17-62"}
    url = extract_lovdata_url(doc, doc_type="LAW")
    assert url == "https://lovdata.no/dokument/NL/lov/2005-06-17-62"


def test_extract_lovdata_url_regulation():
    doc = {"canonical_document_id": "SF/forskrift/2023-02-14-193"}
    url = extract_lovdata_url(doc, doc_type="REGULATION")
    assert url == "https://lovdata.no/dokument/SF/forskrift/2023-02-14-193"


def test_extract_lovdata_url_lti_regulation():
    doc = {"dok_id": "LTI/forskrift/2024-12-20-3428"}
    url = extract_lovdata_url(doc, doc_type="REGULATION")
    assert "forskrift/2024-12-20-3428" in url


def test_extract_lovdata_url_explicit_link():
    doc = {
        "dok_id": "NL/lov/1990-01-01-1",
        "lovdata_url": "https://lovdata.no/dokument/custom-url",
    }
    url = extract_lovdata_url(doc, doc_type="LAW")
    assert url == "https://lovdata.no/dokument/custom-url"


def test_extract_missing_dokid():
    url = extract_lovdata_url({}, "LAW")
    assert url is None


def test_probe_url_200_ok():
    mock_session = MagicMock()
    mock_resp = MagicMock()
    mock_resp.status_code = 200
    mock_resp.url = "https://lovdata.no/dokument/NL/lov/2005-06-17-62"
    mock_session.head.return_value = mock_resp

    res = probe_url("https://lovdata.no/dokument/NL/lov/2005-06-17-62", mock_session)
    assert res["status_code"] == 200
    assert res["result"] == "OK"


def test_probe_url_302_redirect():
    mock_session = MagicMock()
    mock_resp = MagicMock()
    mock_resp.status_code = 302
    mock_resp.url = "https://lovdata.no/dokument/NL/lov/2005-06-17-62-final"
    mock_session.head.return_value = mock_resp

    res = probe_url("https://lovdata.no/dokument/NL/lov/2005-old", mock_session)
    assert res["status_code"] == 302
    assert res["final_url"] == "https://lovdata.no/dokument/NL/lov/2005-06-17-62-final"


def test_probe_url_404_not_found():
    mock_session = MagicMock()
    mock_resp = MagicMock()
    mock_resp.status_code = 404
    mock_resp.url = "https://lovdata.no/dokument/broken"
    mock_session.head.return_value = mock_resp

    res = probe_url("https://lovdata.no/dokument/broken", mock_session)
    assert res["status_code"] == 404
    assert res["result"] == "NOT_FOUND"


def test_probe_url_500_server_error():
    mock_session = MagicMock()
    mock_resp = MagicMock()
    mock_resp.status_code = 500
    mock_resp.url = "https://lovdata.no"
    mock_session.head.return_value = mock_resp

    res = probe_url("https://lovdata.no", mock_session)
    assert res["status_code"] == 500
    assert "500" in res["result"]


def test_probe_url_timeout():
    mock_session = MagicMock()
    mock_session.head.side_effect = requests.Timeout("Connection timed out")

    with patch("time.sleep"):
        res = probe_url("https://lovdata.no/timeout", mock_session, timeout=1.0)
    assert res["status_code"] == 408
    assert "Timeout" in res["result"]


def test_probe_url_connection_error():
    mock_session = MagicMock()
    mock_session.head.side_effect = requests.ConnectionError("Failed to resolve host")

    with patch("time.sleep"):
        res = probe_url("https://lovdata.no/offline", mock_session, timeout=1.0)
    assert res["status_code"] == 0
    assert "ConnectionError" in res["result"]


def test_probe_url_empty():
    mock_session = MagicMock()
    res = probe_url("", mock_session)
    assert res["result"] == "NO_URL"
    assert res["status_code"] is None


def test_audit_empty_documents():
    validator = LovdataURLValidator()
    records = validator.audit_documents([])
    assert records == []


def test_audit_documents_quarantine_flow_and_export_validation(tmp_path):
    validator = LovdataURLValidator(max_workers=2)
    docs = [
        {"dok_id": "NL/lov/valid", "title": "Valid Law", "candidate_domain_ids": ["D12_EMPLOYMENT"]},
        {"dok_id": "NL/lov/broken", "title": "Broken Law", "candidate_domain_ids": ["D12_EMPLOYMENT"]},
    ]

    def mock_probe(url, session, timeout=12.0):
        if "valid" in url:
            return {"status_code": 200, "result": "OK", "final_url": None}
        return {"status_code": 404, "result": "NOT_FOUND", "final_url": None}

    with patch("ingestion.validation.url_validator.probe_url", side_effect=mock_probe):
        records = validator.audit_documents(docs)

    assert len(records) == 2
    valid_record = next(r for r in records if "valid" in r.xapi_id)
    broken_record = next(r for r in records if "broken" in r.xapi_id)

    # 1. Type validation
    assert isinstance(valid_record, URLAuditRecord)
    assert isinstance(broken_record, URLAuditRecord)

    # 2. Status category validation
    assert valid_record.is_valid is True
    assert valid_record.lovdata_status == 200
    assert valid_record.status_category == "FULLY_WORKING"

    assert broken_record.is_valid is False
    assert broken_record.lovdata_status == 404
    assert broken_record.status_category == "XAPI_DATA_LOVDATA_404"

    # 3. Export validation
    excel_path, json_path = validator.export_reports(
        records=records,
        output_dir=tmp_path,
        file_basename="test_audit_report",
    )
    assert excel_path.exists()
    assert json_path.exists()

    # 4. Verify Exported JSON contents
    data = json.loads(json_path.read_text(encoding="utf-8"))
    assert data["total_audited"] == 2
    assert len(data["quarantined_documents"]) == 1
    assert data["quarantined_documents"][0]["xapi_id"] == "NL/lov/broken"
    assert len(data["all_records"]) == 2

    # 5. Verify Exported Excel contents with openpyxl
    wb = openpyxl.load_workbook(excel_path)
    assert "xAPI_Has_Data_Lovdata_404 Error" in wb.sheetnames
    assert "Total Laws and Regulations" in wb.sheetnames
    sheet = wb["Total Laws and Regulations"]
    assert sheet.max_row >= 3  # Header + 2 data rows
    headers = [cell.value for cell in sheet[1]]
    assert "xapi_id" in headers
    assert "lovdata_status" in headers
    assert "status_category" in headers



def test_audit_mixed_batch_concurrency():
    validator = LovdataURLValidator(max_workers=4)
    docs = [
        {"dok_id": "NL/lov/200", "title": "Law 200", "candidate_domain_ids": ["D04_CONTRACT"]},
        {"dok_id": "NL/lov/302", "title": "Law 302", "candidate_domain_ids": ["D04_CONTRACT"]},
        {"dok_id": "NL/lov/404", "title": "Law 404", "candidate_domain_ids": ["D04_CONTRACT"]},
        {"dok_id": "NL/lov/timeout", "title": "Law Timeout", "candidate_domain_ids": ["D04_CONTRACT"]},
        {"dok_id": "", "title": "Missing ID", "candidate_domain_ids": ["D04_CONTRACT"]},
    ]

    def mock_probe(url, session, timeout=12.0):
        if "200" in url:
            return {"status_code": 200, "result": "OK", "final_url": None}
        elif "302" in url:
            return {"status_code": 302, "result": "HTTP_302", "final_url": "https://lovdata.no/target"}
        elif "404" in url:
            return {"status_code": 404, "result": "NOT_FOUND", "final_url": None}
        elif "timeout" in url:
            return {"status_code": 408, "result": "ERROR: Timeout", "final_url": None}
        return {"status_code": None, "result": "NO_URL", "final_url": None}

    with patch("ingestion.validation.url_validator.probe_url", side_effect=mock_probe):
        records = validator.audit_documents(docs)

    status_map = {r.title: r.lovdata_status for r in records}
    assert status_map.get("Law 200") == 200
    assert status_map.get("Law 302") == 302
    assert status_map.get("Law 404") == 404
    assert status_map.get("Law Timeout") == 408
    assert status_map.get("Missing ID") in (None, "ERROR")

