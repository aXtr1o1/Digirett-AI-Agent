from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterator, List

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass(slots=True)
class ExcelRow:
    file_name: str
    domain_name: str
    subdomain_name: str
    b2b_b2c: str
    jurisdiction: str
    relationship_type: str
    tier: str
    source_link: str
    legal_validation: str
    canonical_subdomain: str
    legal_notes: str
    sheet_name: str
    row_number: int


HEADER_ALIASES: dict[str, str] = {
    "File Name": "file_name",
    "Domain Name": "domain_name",
    "Subdomain Name": "subdomain_name",
    "B2B / B2C": "b2b_b2c",
    "Jurisdiction": "jurisdiction",
    "Relationship Type": "relationship_type",
    "Tier": "tier",
    "Source Link": "source_link",
    "Legal_Validation": "legal_validation",
    "Canonical_Subdomain": "canonical_subdomain",
    "Legal_Notes": "legal_notes",
}

REQUIRED_HEADERS = {
    "file_name",
    "domain_name",
    "source_link",
    "legal_validation",
    "canonical_subdomain",
}


class ExcelLoader:
    def __init__(self, workbook_path: str | Path) -> None:
        self.workbook_path = Path(workbook_path)

    def load(self) -> List[ExcelRow]:
        return list(self.iter_rows())

    def iter_rows(self) -> Iterator[ExcelRow]:
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        workbook = load_workbook(
            filename=self.workbook_path,
            read_only=True,
            data_only=True,
        )

        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                header_map = self._build_header_map(worksheet)

                if not header_map:
                    logger.warning("Skipping sheet with no usable header: %s", sheet_name)
                    continue

                missing_required = REQUIRED_HEADERS - set(header_map.values())
                if missing_required:
                    logger.warning(
                        "Sheet missing required headers | sheet=%s | missing=%s",
                        sheet_name,
                        sorted(missing_required),
                    )

                for row_number, row_values in enumerate(
                    worksheet.iter_rows(min_row=2, values_only=True),
                    start=2,
                ):
                    mapped = self._map_row(row_values, header_map)
                    if self._row_is_empty(mapped):
                        continue

                    payload = {
                        "file_name": mapped.get("file_name", ""),
                        "domain_name": mapped.get("domain_name", ""),
                        "subdomain_name": mapped.get("subdomain_name", ""),
                        "b2b_b2c": mapped.get("b2b_b2c", ""),
                        "jurisdiction": mapped.get("jurisdiction", ""),
                        "relationship_type": mapped.get("relationship_type", ""),
                        "tier": mapped.get("tier", ""),
                        "source_link": mapped.get("source_link", ""),
                        "legal_validation": mapped.get("legal_validation", ""),
                        "canonical_subdomain": mapped.get("canonical_subdomain", ""),
                        "legal_notes": mapped.get("legal_notes", ""),
                        "sheet_name": sheet_name,
                        "row_number": row_number,
                    }

                    try:
                        yield ExcelRow(**payload)
                    except Exception as exc:
                        logger.warning(
                            "Skipping invalid row | sheet=%s | row=%s | error=%s",
                            sheet_name,
                            row_number,
                            exc,
                        )
                        continue

        finally:
            workbook.close()

    def _build_header_map(self, worksheet) -> Dict[int, str]:
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if not header_row:
            return {}

        header_map: Dict[int, str] = {}
        for index, raw_header in enumerate(header_row):
            header = self._clean_cell(raw_header)
            if not header:
                continue

            field_name = HEADER_ALIASES.get(header)
            if field_name:
                header_map[index] = field_name

        return header_map

    def _map_row(self, row_values: tuple[Any, ...], header_map: Dict[int, str]) -> Dict[str, str]:
        mapped: Dict[str, str] = {}
        for col_index, field_name in header_map.items():
            value = row_values[col_index] if col_index < len(row_values) else None
            mapped[field_name] = self._clean_cell(value)
        return mapped

    @staticmethod
    def _row_is_empty(mapped: Dict[str, str]) -> bool:
        return not any(value for value in mapped.values())

    @staticmethod
    def _clean_cell(value: Any) -> str:
        if value is None:
            return ""

        try:
            if isinstance(value, float) and value != value:
                return ""
        except Exception:
            pass

        text = str(value).strip()
        if text.lower() in {"", "nan", "none", "null"}:
            return ""

        return text


def load_validated_workbook(workbook_path: str | Path) -> List[ExcelRow]:
    return ExcelLoader(workbook_path).load()


def iter_validated_workbook_rows(workbook_path: str | Path):
    return ExcelLoader(workbook_path).iter_rows()