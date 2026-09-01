
from __future__ import annotations

import json
import logging
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent.parent.parent
REPORTS_DIR = BASE_DIR / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class ReviewItem:
    canonical_document_id: str
    document_type: str
    document_title: str
    section_number: str
    heading: str
    chapter: str
    domain_id: str
    subdomain_id: str
    subdomain_name: str = ""
    review_category: str = "MAPPED"  # "MAPPED" | "NO_SUBDOMAIN" | "UNMAPPED_DOMAIN" | "AI_REVIEW_REQUIRED"
    confidence: float = 1.0
    reason: str = ""
    sample_text: str = ""
    source_url: str = ""
    vdb_eligible: bool = True
    batch_index: Optional[int] = None
    updated_at: str = field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

    @property
    def unique_key(self) -> str:
        sec = self.section_number or self.heading or "sec"
        return f"{self.canonical_document_id}__{sec}__{self.domain_id}"


class TaxonomyReviewExporter:
    """Exports and incrementally appends taxonomy mapping results to a persistent master Excel report."""

    MASTER_FILE_BASENAME = "taxonomy_mapping_master"

    def __init__(self, output_dir: Optional[Union[str, Path]] = None):
        self.output_dir = Path(output_dir) if output_dir else REPORTS_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.master_json_path = self.output_dir / f"{self.MASTER_FILE_BASENAME}.json"
        self.master_excel_path = self.output_dir / f"{self.MASTER_FILE_BASENAME}.xlsx"
        self.latest_excel_path = self.output_dir / "taxonomy_review_report_latest.xlsx"

    def load_master_registry(self) -> Dict[str, Dict[str, Any]]:
        """Loads existing master mapping items from disk registry if present."""
        if self.master_json_path.exists():
            try:
                data = json.loads(self.master_json_path.read_text(encoding="utf-8"))
                items_list = data.get("items", [])
                registry = {}
                for it in items_list:
                    k = f"{it.get('canonical_document_id')}__{it.get('section_number') or it.get('heading')}__{it.get('domain_id')}"
                    registry[k] = it
                return registry
            except Exception as exc:
                logger.warning("Could not read master taxonomy json registry: %s", exc)
        return {}

    def append_batch_items(
        self,
        batch_items: List[ReviewItem],
        batch_index: Optional[int] = None,
    ) -> Tuple[Path, Path]:
       
        registry = self.load_master_registry()

        for item in batch_items:
            if batch_index is not None:
                item.batch_index = batch_index
            item_dict = asdict(item)
            registry[item.unique_key] = item_dict

        # Convert registry back to list of ReviewItem objects
        merged_items = []
        for v in registry.values():
            merged_items.append(
                ReviewItem(
                    canonical_document_id=v.get("canonical_document_id", "UNKNOWN"),
                    document_type=v.get("document_type", "LAW"),
                    document_title=v.get("document_title", "Untitled"),
                    section_number=v.get("section_number", ""),
                    heading=v.get("heading", ""),
                    chapter=v.get("chapter", ""),
                    domain_id=v.get("domain_id", "UNMAPPED"),
                    subdomain_id=v.get("subdomain_id", "NO_SUBDOMAIN"),
                    subdomain_name=v.get("subdomain_name", ""),
                    review_category=v.get("review_category", "MAPPED"),
                    confidence=float(v.get("confidence", 1.0)),
                    reason=v.get("reason", ""),
                    sample_text=v.get("sample_text", ""),
                    source_url=v.get("source_url", ""),
                    vdb_eligible=bool(v.get("vdb_eligible", True)),
                    batch_index=v.get("batch_index"),
                    updated_at=v.get("updated_at", datetime.now(timezone.utc).isoformat()),
                )
            )

        logger.info(
            "Appending Batch %s (%d items) -> Master Registry now has %d total sections.",
            batch_index if batch_index is not None else "-",
            len(batch_items),
            len(merged_items),
        )

        return self.export_excel_report(merged_items, file_basename=self.MASTER_FILE_BASENAME)

    def export_excel_report(
        self,
        items: List[ReviewItem],
        file_basename: Optional[str] = None,
    ) -> Tuple[Path, Path]:
        """
        Exports all items into a styled 6-sheet master Excel workbook and companion JSON file.
        """
        base_name = file_basename or self.MASTER_FILE_BASENAME
        excel_path = self.output_dir / f"{base_name}.xlsx"
        json_path = self.output_dir / f"{base_name}.json"

        # Categorize
        mapped_items = [i for i in items if i.review_category == "MAPPED" or (i.vdb_eligible and i.subdomain_id not in (None, "", "NO_SUBDOMAIN"))]
        no_subdomain_items = [i for i in items if i.review_category == "NO_SUBDOMAIN" or (not i.vdb_eligible and i.domain_id != "UNMAPPED")]
        unmapped_domain_items = [i for i in items if i.review_category == "UNMAPPED_DOMAIN" or i.domain_id in ("UNMAPPED", "OUT_OF_SCOPE")]
        ai_review_items = [i for i in items if i.review_category == "AI_REVIEW_REQUIRED"]

        # Domain breakdown counts
        domain_counts: Dict[str, Dict[str, int]] = {}
        for it in items:
            dom = it.domain_id or "UNMAPPED"
            if dom not in domain_counts:
                domain_counts[dom] = {"total": 0, "mapped": 0, "no_subdomain": 0, "unmapped": 0}
            domain_counts[dom]["total"] += 1
            if it.vdb_eligible and it.subdomain_id not in (None, "", "NO_SUBDOMAIN"):
                domain_counts[dom]["mapped"] += 1
            elif dom == "UNMAPPED":
                domain_counts[dom]["unmapped"] += 1
            else:
                domain_counts[dom]["no_subdomain"] += 1

        # 1. Save JSON master payload
        json_payload = {
            "last_updated_at": datetime.now(timezone.utc).isoformat(),
            "total_sections_tracked": len(items),
            "mapped_in_vdb_count": len(mapped_items),
            "no_subdomain_count": len(no_subdomain_items),
            "unmapped_domain_count": len(unmapped_domain_items),
            "ai_review_required_count": len(ai_review_items),
            "domain_breakdown": domain_counts,
            "items": [asdict(i) for i in items],
        }
        json_path.write_text(json.dumps(json_payload, indent=2, ensure_ascii=False), encoding="utf-8")

        # 2. Build Styled Excel Workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Styling definitions
        teal_fill = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
        navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        green_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        amber_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        red_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")

        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="E0E0E0"),
            right=Side(style="thin", color="E0E0E0"),
            top=Side(style="thin", color="E0E0E0"),
            bottom=Side(style="thin", color="E0E0E0"),
        )
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        align_center = Alignment(horizontal="center", vertical="center")

        headers = [
            "Doc ID",
            "Doc Type",
            "Title",
            "Section",
            "Heading",
            "Chapter",
            "Domain ID",
            "Subdomain ID",
            "Subdomain Name",
            "VDB Eligible",
            "Category",
            "Confidence",
            "Reason / Notes",
            "Sample Text",
            "Source URL",
        ]

        def _populate_sheet(ws, sheet_items: List[ReviewItem], header_fill: PatternFill):
            ws.views.sheetView[0].showGridLines = True
            for col_idx, h_text in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = align_center
                cell.border = thin_border
            ws.row_dimensions[1].height = 28

            for row_idx, item in enumerate(sheet_items, start=2):
                ws.row_dimensions[row_idx].height = 20
                row_values = [
                    item.canonical_document_id,
                    item.document_type,
                    item.document_title,
                    item.section_number,
                    item.heading,
                    item.chapter,
                    item.domain_id,
                    item.subdomain_id or "NO_SUBDOMAIN",
                    item.subdomain_name or "",
                    "YES" if item.vdb_eligible else "NO",
                    item.review_category,
                    f"{item.confidence:.2f}" if item.confidence is not None else "0.00",
                    item.reason,
                    item.sample_text[:400] if item.sample_text else "",
                    item.source_url,
                ]
                for col_idx, val in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    cell.alignment = align_center if col_idx in (2, 4, 7, 8, 10, 11, 12) else align_left

            col_widths = {
                1: 26, 2: 12, 3: 35, 4: 12, 5: 28, 6: 18,
                7: 18, 8: 16, 9: 25, 10: 14, 11: 20, 12: 12, 13: 35, 14: 45, 15: 35
            }
            for col_idx, width in col_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = width

        # -------------------------------------------------------------
        # Sheet 1: Summary & Overview
        # -------------------------------------------------------------
        ws_summary = wb.create_sheet(title="Summary & Overview")
        ws_summary.views.sheetView[0].showGridLines = True
        ws_summary.column_dimensions["A"].width = 38
        ws_summary.column_dimensions["B"].width = 20

        summary_rows = [
            ("DigiRett Master Taxonomy Mapping & Review", ""),
            ("Last Updated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Sections Tracked Across Batches", len(items)),
            ("1. Mapped into Vector DB (Valid Domain + Subdomain)", len(mapped_items)),
            ("2. No Subdomain Match (General Provisions Excluded from VDB)", len(no_subdomain_items)),
            ("3. Unmapped Domain (No Legal Domain)", len(unmapped_domain_items)),
            ("4. AI Review Required (Low Confidence / Confusers)", len(ai_review_items)),
        ]

        ws_summary.cell(row=1, column=1, value="Metric / Section Category").fill = teal_fill
        ws_summary.cell(row=1, column=1).font = header_font
        ws_summary.cell(row=1, column=2, value="Count").fill = teal_fill
        ws_summary.cell(row=1, column=2).font = header_font
        ws_summary.row_dimensions[1].height = 28

        for r_idx, (k, v) in enumerate(summary_rows, start=2):
            ws_summary.row_dimensions[r_idx].height = 22
            c1 = ws_summary.cell(row=r_idx, column=1, value=k)
            c2 = ws_summary.cell(row=r_idx, column=2, value=v)
            c1.font = bold_font if "Total" in k or r_idx == 2 else data_font
            c2.font = bold_font if "Total" in k else data_font
            c1.border = thin_border
            c2.border = thin_border

        # Domain Breakdown Table on Sheet 1
        d_start = len(summary_rows) + 3
        ws_summary.cell(row=d_start, column=1, value="Domain ID").fill = teal_fill
        ws_summary.cell(row=d_start, column=1).font = header_font
        ws_summary.cell(row=d_start, column=2, value="Total").fill = teal_fill
        ws_summary.cell(row=d_start, column=2).font = header_font
        ws_summary.cell(row=d_start, column=3, value="In VDB (Mapped)").fill = teal_fill
        ws_summary.cell(row=d_start, column=3).font = header_font
        ws_summary.cell(row=d_start, column=4, value="No Subdomain").fill = teal_fill
        ws_summary.cell(row=d_start, column=4).font = header_font
        ws_summary.column_dimensions["C"].width = 20
        ws_summary.column_dimensions["D"].width = 20

        for r_offset, (dom, stats) in enumerate(sorted(domain_counts.items()), start=1):
            row_num = d_start + r_offset
            ws_summary.row_dimensions[row_num].height = 20
            ws_summary.cell(row=row_num, column=1, value=dom).font = data_font
            ws_summary.cell(row=row_num, column=2, value=stats["total"]).font = data_font
            ws_summary.cell(row=row_num, column=3, value=stats["mapped"]).font = data_font
            ws_summary.cell(row=row_num, column=4, value=stats["no_subdomain"]).font = data_font
            for c_idx in range(1, 5):
                ws_summary.cell(row=row_num, column=c_idx).border = thin_border

        # -------------------------------------------------------------
        # Sheet 2: All Sections & Mappings
        # -------------------------------------------------------------
        ws_all = wb.create_sheet(title="All Sections & Mappings")
        _populate_sheet(ws_all, items, navy_fill)

        # -------------------------------------------------------------
        # Sheet 3: Mapped Sections (In VDB)
        # -------------------------------------------------------------
        ws_mapped = wb.create_sheet(title="Mapped Sections (In VDB)")
        _populate_sheet(ws_mapped, mapped_items, green_fill)

        # -------------------------------------------------------------
        # Sheet 4: No Subdomain (Excluded VDB)
        # -------------------------------------------------------------
        ws_nosub = wb.create_sheet(title="No Subdomain (Excluded VDB)")
        _populate_sheet(ws_nosub, no_subdomain_items, amber_fill)

        # -------------------------------------------------------------
        # Sheet 5: Unmapped Domain
        # -------------------------------------------------------------
        ws_unmap = wb.create_sheet(title="Unmapped Domain")
        _populate_sheet(ws_unmap, unmapped_domain_items, red_fill)

        # -------------------------------------------------------------
        # Sheet 6: AI Review Required
        # -------------------------------------------------------------
        ws_aireview = wb.create_sheet(title="AI Review Required")
        _populate_sheet(ws_aireview, ai_review_items, purple_fill)

        # Save workbooks
        wb.save(excel_path)
        wb.save(self.latest_excel_path)

        logger.info(
            "Master Excel Workbook saved: %d total sections -> %s",
            len(items),
            excel_path.name,
        )
        return excel_path, json_path

    @classmethod
    def export_from_supabase_db(cls, supabase_store: Any) -> Tuple[Path, Path]:
        """Fetches all sections from Supabase legal_sections table and populates the master Excel report."""
        if not supabase_store or not supabase_store._ensure_connection() or not supabase_store.supabase:
            raise RuntimeError("Supabase client is not connected.")

        logger.info("Fetching complete section mapping dataset from Supabase legal_sections table...")
        items: List[ReviewItem] = []
        offset = 0
        page_size = 1000

        while True:
            res = (
                supabase_store.supabase.table("legal_sections")
                .select(
                    "legal_document_id, section_number, section_title, chapter_number, chapter_title, source_section_key, source_section_url, classification, vdb_eligible, vdb_status, legal_documents(canonical_document_id, document_type, title, primary_domain_id)"
                )
                .range(offset, offset + page_size - 1)
                .execute()
            )
            rows = res.data or []
            for row in rows:
                parent_doc = row.get("legal_documents") or {}
                doc_id = parent_doc.get("canonical_document_id") or "UNKNOWN"
                doc_type = parent_doc.get("document_type") or "LAW"
                doc_title = parent_doc.get("title") or doc_id
                sec_num = row.get("section_number") or ""
                heading = row.get("section_title") or sec_num or row.get("source_section_key", "")
                chap = f"{row.get('chapter_number', '') or ''} {row.get('chapter_title', '') or ''}".strip()

                doms = row.get("classification") or []
                if isinstance(doms, list):
                    primary_dom = parent_doc.get("primary_domain_id") or (doms[0] if doms else "UNMAPPED")
                else:
                    primary_dom = parent_doc.get("primary_domain_id") or "UNMAPPED"

                vdb_eligible = bool(row.get("vdb_eligible", False))
                vdb_status = str(row.get("vdb_status", "NOT_ELIGIBLE"))

                if vdb_eligible and vdb_status == "SYNCED":
                    cat = "MAPPED"
                    reason = "Valid Domain and Subdomain mapped"
                    primary_sub = "MAPPED"
                elif not doms or primary_dom in ("UNMAPPED", "OUT_OF_SCOPE"):
                    cat = "UNMAPPED_DOMAIN"
                    reason = "No candidate legal domain mapped"
                    primary_sub = "UNMAPPED"
                else:
                    cat = "NO_SUBDOMAIN"
                    reason = "General domain provision or no specific subdomain matched (Excluded from VDB)"
                    primary_sub = "NO_SUBDOMAIN"

                src_url = row.get("source_section_url") or ""

                items.append(
                    ReviewItem(
                        canonical_document_id=doc_id,
                        document_type=doc_type,
                        document_title=doc_title,
                        section_number=sec_num,
                        heading=heading,
                        chapter=chap,
                        domain_id=primary_dom,
                        subdomain_id=primary_sub,
                        review_category=cat,
                        confidence=1.0 if cat == "MAPPED" else 0.0,
                        reason=reason,
                        sample_text="",
                        source_url=src_url,
                        vdb_eligible=vdb_eligible,
                    )
                )

            if len(rows) < page_size:
                break
            offset += page_size

        exporter = cls()
        return exporter.export_excel_report(items, file_basename=exporter.MASTER_FILE_BASENAME)
