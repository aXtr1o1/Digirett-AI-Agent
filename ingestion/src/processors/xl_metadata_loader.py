"""
xl_metadata_loader.py
======================
Loads Excel dataset files for per-domain pipeline processing.

Key difference from the old loader:
  - load_all_xl_files() returns a LIST of (domain_name, url_map) tuples,
    one entry per XL file, preserving per-file identity.
  - Each url_map maps { normalised_url: XlFileMetadata } containing only
    the sub_domain_name and source_type from THAT specific XL file.
  - The pipeline processes one XL file at a time, so domain accumulation
    happens progressively via Supabase/Milvus appends — not upfront.

XL column-name variants handled:
    URL column   : Lovdata_URL, URL
    subdomain    : Korttittel, Navn, Fulltittel  (first found wins per file)
    source_type  : Type
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Per-row metadata from a single XL file
# ---------------------------------------------------------------------------

class XlFileMetadata:
    """Metadata for one URL entry from one specific XL file."""

    __slots__ = ("sub_domain_name", "source_type")

    def __init__(self, sub_domain_name: str = "", source_type: str = ""):
        self.sub_domain_name = sub_domain_name
        self.source_type     = source_type

    def __repr__(self) -> str:
        return f"XlFileMetadata(sub={self.sub_domain_name!r}, type={self.source_type!r})"


# { normalised_url: XlFileMetadata }
XlFileMap = Dict[str, XlFileMetadata]

# List of (domain_name, XlFileMap) — one entry per XL file
XlDomainList = List[Tuple[str, XlFileMap]]


# ---------------------------------------------------------------------------
# URL normalisation helpers
# ---------------------------------------------------------------------------

def _normalise(url: str) -> str:
    return str(url).strip().rstrip("/")


def _both_forms(url: str) -> List[str]:
    """Return both short and long Lovdata URL forms so lookups always hit."""
    variants = [url]
    if "/lovdata.no/lov/" in url:
        variants.append(url.replace("/lovdata.no/lov/", "/lovdata.no/dokument/NL/lov/"))
    elif "/lovdata.no/dokument/NL/lov/" in url:
        variants.append(url.replace("/lovdata.no/dokument/NL/lov/", "/lovdata.no/lov/"))
    if "/lovdata.no/forskrift/" in url:
        variants.append(url.replace("/lovdata.no/forskrift/", "/lovdata.no/dokument/SF/forskrift/"))
    elif "/lovdata.no/dokument/SF/forskrift/" in url:
        variants.append(url.replace("/lovdata.no/dokument/SF/forskrift/", "/lovdata.no/forskrift/"))
    return list(dict.fromkeys(variants))


def _find_col(headers_lower: List[str], candidates: List[str]) -> Optional[int]:
    for candidate in candidates:
        for i, h in enumerate(headers_lower):
            if h == candidate:
                return i
    return None


# ---------------------------------------------------------------------------
# Load a single XL file → XlFileMap
# ---------------------------------------------------------------------------

def load_xl_single_file(xl_path: Path) -> Tuple[str, XlFileMap]:
    """
    Load one Excel file.
    Returns (domain_name, { normalised_url: XlFileMetadata }).
    domain_name = file stem (e.g. "Selskapsrett").
    """
    import openpyxl

    domain_name = xl_path.stem
    url_map: XlFileMap = {}

    try:
        wb = openpyxl.load_workbook(xl_path, read_only=True, data_only=True)
        ws = wb.active

        raw_headers = list(next(ws.iter_rows(max_row=1, values_only=True), []))
        if not raw_headers:
            logger.warning(f"  ⚠️  No headers: {xl_path.name}")
            wb.close()
            return domain_name, url_map

        headers_lower = [str(h).strip().lower() if h else "" for h in raw_headers]

        url_col  = _find_col(headers_lower, ["lovdata_url", "url"])
        if url_col is None:
            logger.warning(f"  ⚠️  No URL column in {xl_path.name} — skipping")
            wb.close()
            return domain_name, url_map

        sub_col  = _find_col(headers_lower, ["korttittel", "navn", "fulltittel"])
        type_col = _find_col(headers_lower, ["type"])

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_url = row[url_col] if url_col < len(row) else None
            if not raw_url or not str(raw_url).strip().startswith("http"):
                continue

            base_url = _normalise(str(raw_url))

            sub = (
                str(row[sub_col]).strip()
                if sub_col is not None and sub_col < len(row) and row[sub_col]
                else ""
            )
            src = (
                str(row[type_col]).strip()
                if type_col is not None and type_col < len(row) and row[type_col]
                else ""
            )

            meta = XlFileMetadata(sub_domain_name=sub, source_type=src)

            # Index under both URL forms (first writer wins within same file)
            for variant in _both_forms(base_url):
                if variant not in url_map:
                    url_map[variant] = meta

            count += 1

        wb.close()
        logger.info(
            f"  ✅ {xl_path.name}: {count} rows | "
            f"domain='{domain_name}' | "
            f"sub_col={'yes (' + headers_lower[sub_col] + ')' if sub_col is not None else 'NO'} | "
            f"type_col={'yes' if type_col is not None else 'NO'}"
        )

    except Exception as e:
        logger.error(f"  ❌ Failed to load {xl_path.name}: {e}")

    return domain_name, url_map


# ---------------------------------------------------------------------------
# Load all XL files → ordered list of (domain_name, XlFileMap)
# ---------------------------------------------------------------------------

def load_all_xl_files(xl_folder: str) -> XlDomainList:
    """
    Scan all .xlsx / .xlsm files in xl_folder.
    Returns a list of (domain_name, url_map) tuples — ONE per XL file.
    Order is alphabetical by filename.

    The pipeline iterates this list and processes each domain independently,
    appending metadata to existing rows rather than merging upfront.
    """
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise ImportError("pip install openpyxl")

    folder = Path(xl_folder.replace("\\", "/"))
    if not folder.exists():
        raise FileNotFoundError(f"❌ XL folder not found: {folder.resolve()}")

    xl_files = sorted([
        f for f in (list(folder.glob("*.xlsx")) + list(folder.glob("*.xlsm")))
        if not f.name.startswith("~$")
    ])
    if not xl_files:
        raise FileNotFoundError(f"❌ No Excel files in: {folder.resolve()}")

    logger.info(f"📊 Loading {len(xl_files)} XL domain files from '{xl_folder}'")

    result: XlDomainList = []
    for f in xl_files:
        domain_name, url_map = load_xl_single_file(f)
        if url_map:
            result.append((domain_name, url_map))

    total_urls = sum(len(m) for _, m in result)
    logger.info(f"📊 Loaded {len(result)} domains | {total_urls} total URL entries")
    return result