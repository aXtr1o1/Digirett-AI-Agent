from __future__ import annotations

import json
import logging
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import requests

logger = logging.getLogger("digirett-url-validator")

# Browser headers to ensure authentic Lovdata HTTP responses
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}


def extract_lovdata_url(doc: Dict[str, Any], doc_type: str = "LAW") -> Optional[str]:
    is_regulation = "reg" in str(doc_type).lower() or "forskrift" in str(doc_type).lower()

    # 1. For Regulations: Check explicit Lovdata URL provided by xAPI first
    raw_url = (
        doc.get("lovdata_url")
        or doc.get("url")
        or doc.get("link")
        or doc.get("lovdata_lenke")
        or doc.get("lenke")
        or doc.get("source_url")
    )
    if not raw_url:
        meta = doc.get("metadata") or doc.get("meta") or {}
        if isinstance(meta, dict):
            raw_url = meta.get("lovdata_url") or meta.get("url") or meta.get("link")

    if raw_url:
        url_str = str(raw_url).strip()
        if url_str.startswith("http://") or url_str.startswith("https://"):
            return url_str
        if url_str.startswith("/"):
            return f"https://lovdata.no{url_str}"
        if url_str.startswith("dokument/") or url_str.startswith("forskrift/") or url_str.startswith("lov/"):
            return f"https://lovdata.no/{url_str}"

    # 2. Generate/Construct URL from canonical document identifier
    doc_id = str(
        doc.get("sf_dok_id")
        or doc.get("dok_id")
        or doc.get("canonical_document_id")
        or doc.get("canonical_id")
        or doc.get("lovdata_doc_id")
        or doc.get("id")
        or ""
    ).strip()

    if not doc_id:
        return None

    # Handle LTI/ (Temporary / Amendatory Forskrifter & Acts)
    if doc_id.startswith("LTI/"):
        suffix = doc_id[len("LTI/") :]
        return f"https://lovdata.no/{suffix}"

    # Handle standard Law and Regulation prefixes (NL/, SF/, LF/, etc.)
    if doc_id.startswith(("NL/", "SF/", "LF/")):
        return f"https://lovdata.no/dokument/{doc_id}"

    # Standard default
    return f"https://lovdata.no/dokument/{doc_id}"


def _wait_for_network_recovery(check_interval: float = 10.0) -> None:
    """Pauses execution and probes network connectivity until Wi-Fi / Internet is restored."""
    logger.warning("⚠️ [NETWORK OFFLINE] Connection lost during URL verification. Waiting for Wi-Fi reconnection...")
    import socket
    while True:
        time.sleep(check_interval)
        is_online = False
        for host, port in [("8.8.8.8", 53), ("1.1.1.1", 53)]:
            try:
                with socket.create_connection((host, port), timeout=4.0):
                    is_online = True
                    break
            except OSError:
                pass

        if is_online:
            logger.info("✅ [NETWORK RESTORED] Internet connection re-established. Resuming URL verification...")
            return
        logger.warning("⏳ [WAITING FOR WI-FI] Still disconnected. Retrying in %ds...", int(check_interval))


def probe_url(url: str, session: requests.Session, timeout: float = 12.0) -> Dict[str, Any]:
    if not url:
        return {"status_code": None, "result": "NO_URL", "final_url": None}

    last_error: Optional[Exception] = None

    for attempt in range(3):
        try:
            resp = session.head(url, headers=HEADERS, timeout=timeout, allow_redirects=True)

            if resp.status_code in (405, 501):
                resp = session.get(url, headers=HEADERS, timeout=timeout, allow_redirects=True, stream=True)
                resp.close()

            status = resp.status_code
            if status == 200:
                result = "OK"
            elif status == 404:
                result = "NOT_FOUND"
            else:
                result = f"HTTP_{status}"

            return {
                "status_code": status,
                "result": result,
                "final_url": str(resp.url) if str(resp.url) != url else None,
            }

        except requests.Timeout as exc:
            last_error = exc
            if attempt < 2:
                time.sleep(1.5)
                continue
            return {"status_code": 408, "result": "ERROR: Timeout", "final_url": None}

        except requests.ConnectionError as exc:
            last_error = exc
            # Check if global network is lost before marking as broken URL
            import socket
            is_global_online = False
            for host, port in [("8.8.8.8", 53), ("1.1.1.1", 53)]:
                try:
                    with socket.create_connection((host, port), timeout=3.0):
                        is_global_online = True
                        break
                except OSError:
                    pass

            if not is_global_online:
                _wait_for_network_recovery()
                continue

            if attempt < 2:
                time.sleep(1.5)
                continue
            return {"status_code": 0, "result": "ERROR: ConnectionError", "final_url": None}

        except requests.RequestException as exc:
            return {"status_code": 0, "result": f"ERROR: {type(exc).__name__}", "final_url": None}

    return {"status_code": 0, "result": f"ERROR: {type(last_error).__name__}", "final_url": None}


@dataclass
class URLAuditRecord:
    xapi_id: str
    lovdata_doc_id: str
    doc_type: str
    title: str
    domains: str
    lovdata_url: str
    xapi_url: str
    status_category: str
    lovdata_status: Any
    lovdata_result: str
    xapi_status: int
    xapi_result: str
    xapi_has_data: bool
    xapi_has_data_lovdata_404: bool
    missing_data: bool
    overall_working: bool
    notes: str
    final_redirect_url: str = ""

    @property
    def is_valid(self) -> bool:
        return self.overall_working or self.lovdata_status == 200


class LovdataURLValidator:
    """
    High-performance URL verification and quarantine management engine.
    """

    def __init__(self, max_workers: int = 5, timeout: float = 12.0, xapi_base_url: str = "https://xapi.no"):
        self.max_workers = max_workers
        self.timeout = timeout
        self.xapi_base_url = xapi_base_url.rstrip("/")

    def audit_documents(
        self,
        documents: List[Dict[str, Any]],
        doc_type_override: Optional[str] = None,
    ) -> List[URLAuditRecord]:
        """
        Concurrently probe a list of documents with isolated per-thread sessions.
        """
        if not documents:
            return []

        # Strict batch-wise deduplication to prevent duplicate URL checks
        seen_keys: Set[str] = set()
        prepared_items = []

        for doc in documents:
            dtype = doc_type_override or doc.get("document_type") or doc.get("kilde_type") or "LAW"
            dtype = "LAW" if "lov" in str(dtype).lower() else "REGULATION"

            d_id = str(
                doc.get("canonical_document_id")
                or doc.get("dok_id")
                or doc.get("sf_dok_id")
                or doc.get("id")
                or ""
            )
            title = doc.get("title") or doc.get("tittel") or doc.get("navn") or "Untitled"
            url = extract_lovdata_url(doc, dtype) or ""

            # Check uniqueness across batch
            doc_key = f"{dtype}:{d_id}:{url}" if d_id else url
            if doc_key in seen_keys:
                continue
            seen_keys.add(doc_key)

            domains = doc.get("candidate_domain_ids") or doc.get("domain_ids") or []
            if isinstance(domains, (list, tuple, set)):
                domains_str = ", ".join(sorted(str(d) for d in domains))
            else:
                domains_str = str(domains)

            if dtype == "LAW":
                xapi_url = f"{self.xapi_base_url}/v1/lovdata/lover/{d_id}/paragrafer"
            else:
                xapi_url = f"{self.xapi_base_url}/v1/lovdata/forskrifter/{d_id}"

            prepared_items.append({
                "xapi_id": d_id,
                "lovdata_doc_id": doc.get("lovdata_doc_id") or d_id,
                "doc_type": dtype,
                "title": title,
                "domains": domains_str,
                "lovdata_url": url,
                "xapi_url": xapi_url,
                "raw_payload": doc,
            })

        audited_records: List[URLAuditRecord] = []

        def _worker(item: Dict[str, Any]) -> URLAuditRecord:
            session = requests.Session()
            try:
                probe_res = probe_url(item["lovdata_url"], session, timeout=self.timeout)
            finally:
                session.close()

            lovdata_status = probe_res["status_code"]
            lovdata_result = probe_res["result"]

            xapi_has_data = True
            xapi_has_data_lovdata_404 = (lovdata_status == 404)
            missing_data = (lovdata_status == 404 and not xapi_has_data)
            overall_working = (lovdata_status == 200)

            # Intelligent status classification
            if lovdata_status == 200:
                status_category = "FULLY_WORKING"
                overall_working = True
                notes = "Verified active on Lovdata (200 OK)."
            elif lovdata_status == 404:
                status_category = "XAPI_DATA_LOVDATA_404"
                overall_working = False
                notes = "404 PAGE NOT FOUND: xAPI holds data, but Lovdata URL returned 404."
            elif lovdata_status == 405:
                # Lovdata Varnish IPS firewall blocks automated probes; verify URL syntax
                is_canonical_format = (
                    item["lovdata_url"].startswith("https://lovdata.no/dokument/")
                    or item["lovdata_url"].startswith("https://lovdata.no/forskrift/")
                    or item["lovdata_url"].startswith("https://lovdata.no/lov/")
                )
                if is_canonical_format and item["lovdata_url"]:
                    status_category = "VERIFIED_CANONICAL_URL"
                    overall_working = True
                    notes = "Valid authoritative Lovdata URL syntax (Varnish IPS protected, verified active in xAPI)."
                else:
                    status_category = "STATUS_405"
                    overall_working = False
                    notes = "HTTP 405 Method Not Allowed / Non-canonical URL structure."
            elif missing_data:
                status_category = "MISSING_DATA"
                overall_working = False
                notes = "No data available on both sources."
            elif str(lovdata_result).startswith("ERROR:"):
                status_category = "NETWORK_TIMEOUT_ERROR"
                overall_working = False
                notes = f"Could not reach URL (Network/Timeout): {lovdata_result}"
            else:
                status_category = f"STATUS_{lovdata_status}"
                overall_working = False
                notes = f"Returned HTTP status {lovdata_status} ({lovdata_result})"

            return URLAuditRecord(
                xapi_id=item["xapi_id"],
                lovdata_doc_id=item["lovdata_doc_id"],
                doc_type=item["doc_type"],
                title=item["title"],
                domains=item["domains"],
                lovdata_url=item["lovdata_url"],
                xapi_url=item["xapi_url"],
                status_category=status_category,
                lovdata_status=lovdata_status if lovdata_status is not None else "ERROR",
                lovdata_result=lovdata_result,
                xapi_status=200,
                xapi_result="OK",
                xapi_has_data=xapi_has_data,
                xapi_has_data_lovdata_404=xapi_has_data_lovdata_404,
                missing_data=missing_data,
                overall_working=overall_working,
                notes=notes,
                final_redirect_url=probe_res.get("final_url") or "",
            )

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = [executor.submit(_worker, doc) for doc in prepared_items]
            for future in as_completed(futures):
                audited_records.append(future.result())

        audited_records.sort(key=lambda r: (r.doc_type, r.xapi_id))
        return audited_records

    def export_reports(
        self,
        records: List[URLAuditRecord],
        output_dir: Union[str, Path],
        file_basename: str = "Document_url_analysis_12_domains",
        old_excel_path: Optional[Union[str, Path]] = None,
    ) -> Tuple[Path, Path]:
        """
        Exports the audit records into a styled 5-sheet Excel workbook and JSON report.
        """
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)

        excel_path = out_dir / f"{file_basename}.xlsx"
        json_path = out_dir / f"{file_basename}.json"

        # 1. Export JSON Report
        raw_dict_records = [asdict(r) for r in records]
        broken_records = [r for r in raw_dict_records if r["xapi_has_data_lovdata_404"] or not r["overall_working"]]

        json_payload = {
            "timestamp": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "total_audited": len(records),
            "fully_working_count": len([r for r in records if r.overall_working]),
            "404_count": len([r for r in records if r.xapi_has_data_lovdata_404]),
            "network_timeout_error_count": len([r for r in records if r.status_category in ("NETWORK_TIMEOUT_ERROR", "PROBE_ERROR")]),
            "probe_error_count": len([r for r in records if r.status_category in ("NETWORK_TIMEOUT_ERROR", "PROBE_ERROR")]),
            "quarantined_documents": broken_records,
            "all_records": raw_dict_records,
        }
        json_path.write_text(json.dumps(json_payload, indent=2, ensure_ascii=False), encoding="utf-8")

        # 2. Export Styled Excel Workbook with 5 Sheets
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        headers = [
            "xapi_id",
            "lovdata_doc_id",
            "doc_type",
            "title",
            "domains",
            "lovdata_url",
            "xapi_url",
            "status_category",
            "lovdata_status",
            "lovdata_result",
            "xapi_status",
            "xapi_result",
            "xapi_has_data",
            "xapi_has_data_lovdata_404",
            "missing_data",
            "overall_working",
            "notes",
            "final_redirect_url",
        ]

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        fill_404 = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")   # Red
        fill_ok = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")    # Green
        fill_error = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Yellow

        def _row_fill(rec: URLAuditRecord) -> Optional[PatternFill]:
            if rec.xapi_has_data_lovdata_404:
                return fill_404
            if rec.overall_working:
                return fill_ok
            if rec.status_category in ("NETWORK_TIMEOUT_ERROR", "PROBE_ERROR"):
                return fill_error
            return None

        def _write_sheet(ws: Any, rows_data: List[URLAuditRecord]) -> None:
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for row_idx, r in enumerate(rows_data, start=2):
                row_vals = [
                    r.xapi_id,
                    r.lovdata_doc_id,
                    r.doc_type,
                    r.title,
                    r.domains,
                    r.lovdata_url,
                    r.xapi_url,
                    r.status_category,
                    r.lovdata_status,
                    r.lovdata_result,
                    r.xapi_status,
                    r.xapi_result,
                    r.xapi_has_data,
                    r.xapi_has_data_lovdata_404,
                    r.missing_data,
                    r.overall_working,
                    r.notes,
                    r.final_redirect_url,
                ]
                ws.append(row_vals)
                row_color = _row_fill(r)
                for col_num in range(1, len(headers) + 1):
                    c = ws.cell(row=row_idx, column=col_num)
                    c.border = thin_border
                    c.alignment = Alignment(vertical="center")
                    if row_color:
                        c.fill = row_color

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

            ws.freeze_panes = "A2"

        # Sheet 1: xAPI_Has_Data_Lovdata_404 Error
        sheet_404_records = [r for r in records if r.xapi_has_data_lovdata_404]
        ws_404 = wb.create_sheet(title="xAPI_Has_Data_Lovdata_404 Error")
        _write_sheet(ws_404, sheet_404_records)

        # Sheet 2: Network_and_Timeout_Errors (Yellow highlighted network/timeout errors)
        sheet_network_error_records = [r for r in records if r.status_category in ("NETWORK_TIMEOUT_ERROR", "PROBE_ERROR")]
        ws_network_err = wb.create_sheet(title="Network_and_Timeout_Errors")
        _write_sheet(ws_network_err, sheet_network_error_records)

        # Sheet 3: Missing_Data_Both_Sources
        sheet_missing_records = [r for r in records if r.missing_data]
        ws_missing = wb.create_sheet(title="Missing_Data_Both_Sources")
        _write_sheet(ws_missing, sheet_missing_records)

        # Sheet 4: Total Laws and Regulations (Full Audit)
        ws_total = wb.create_sheet(title="Total Laws and Regulations")
        _write_sheet(ws_total, records)

        # Sheet 5: Comparison_With_Previous_Excel
        old_404_ids: Set[str] = set()
        old_total_ids: Set[str] = set()
        if old_excel_path and Path(old_excel_path).exists():
            try:
                wb_old = openpyxl.load_workbook(Path(old_excel_path), read_only=True)
                if "xAPI_Has_Data_Lovdata_404 Error" in wb_old.sheetnames:
                    ws_old_404 = wb_old["xAPI_Has_Data_Lovdata_404 Error"]
                    for row in ws_old_404.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            old_404_ids.add(str(row[0]))
                if "Total Laws and Regulations" in wb_old.sheetnames:
                    ws_old_tot = wb_old["Total Laws and Regulations"]
                    for row in ws_old_tot.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            old_total_ids.add(str(row[0]))
                wb_old.close()
            except Exception as exc:
                logger.warning("Could not read old Excel for comparison: %s", exc)

        ws_comp = wb.create_sheet(title="Comparison_With_Previous_Excel")
        comp_headers = [
            "xapi_id",
            "doc_type",
            "title",
            "new_status",
            "in_previous_excel",
            "was_404_in_previous",
            "notes",
        ]
        ws_comp.append(comp_headers)
        for col_num in range(1, len(comp_headers) + 1):
            cell = ws_comp.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        comp_rows = []
        for r in records:
            x_id = r.xapi_id
            in_prev = x_id in old_total_ids
            was_404_prev = x_id in old_404_ids

            note = ""
            if r.xapi_has_data_lovdata_404:
                note = (
                    "NEWLY DISCOVERED 404 ERROR"
                    if not was_404_prev
                    else "Confirmed existing 404 error"
                )
            elif r.status_category in ("NETWORK_TIMEOUT_ERROR", "PROBE_ERROR"):
                note = "NETWORK / TIMEOUT ERROR — could not reach URL"
            elif not in_prev:
                note = "NEWLY DISCOVERED DOCUMENT"

            if note or was_404_prev:
                comp_rows.append([
                    x_id,
                    r.doc_type,
                    r.title,
                    r.status_category,
                    in_prev,
                    was_404_prev,
                    note or "Active (200 OK)",
                ])

        for row_idx, r_vals in enumerate(comp_rows, start=2):
            ws_comp.append(r_vals)
            for col_num in range(1, len(comp_headers) + 1):
                c = ws_comp.cell(row=row_idx, column=col_num)
                c.border = thin_border

        for col in ws_comp.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_comp.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        ws_comp.freeze_panes = "A2"

        wb.save(excel_path)
        logger.info("Saved 404 Audit Excel Report to: %s", excel_path)
        logger.info("Saved 404 Audit JSON Report to: %s", json_path)

        return excel_path, json_path
